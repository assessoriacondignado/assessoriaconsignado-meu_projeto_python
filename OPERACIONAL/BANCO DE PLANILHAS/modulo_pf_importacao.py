import streamlit as st
import pandas as pd
import io
import os
import re
import openpyxl
from datetime import datetime
import modulo_pf_cadastro as pf_core

# --- CONFIGURAÇÕES DE DIRETÓRIO ---
BASE_DIR_IMPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ARQUIVO IMPORTAÇÕES")
if not os.path.exists(BASE_DIR_IMPORTS):
    os.makedirs(BASE_DIR_IMPORTS)

def get_table_columns(table_name):
    conn = pf_core.get_conn()
    cols = []
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = '{table_name}' AND table_schema = 'banco_pf'")
            cols = cur.fetchall()
            conn.close()
        except: pass
    return cols

def validar_planilha_estrita(caminho_arquivo):
    """
    Valida se as linhas 1 e 2 possuem formatação 'Geral'.
    Retorna: (bool, msg_erro)
    """
    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=False)
        ws = wb.active
        
        for col_index, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=2), start=1):
            cabecalho_val = col_cells[0].value
            nome_coluna = str(cabecalho_val) if cabecalho_val else openpyxl.utils.get_column_letter(col_index)
            
            for cell in col_cells:
                fmt = str(cell.number_format).lower()
                if fmt == 'general':
                    return False, (
                        f"⛔ **Bloqueio de Importação**: A coluna **'{nome_coluna}'** está com formatação **'Geral'** na linha {cell.row}. "
                        "Para garantir a integridade, converta todas as colunas para **TEXTO** ou **NÚMERO** no Excel antes de importar."
                    )
        return True, None

    except Exception as e:
        return False, f"Erro ao validar planilha: {str(e)}"

# --- FUNÇÃO HELPER DE LIMPEZA CPF (REGRAS 2.1) ---
def limpar_cpf_regra_importacao(valor):
    """
    Aplica as regras 2.1:
    - Remove espaços (2.1.3)
    - Remove letras/especiais (2.1.4)
    - Remove zeros à esquerda para armazenamento (2.1.1)
    """
    if pd.isna(valor) or valor == "":
        return ""
    
    # 1. Converte para string e remove espaços (2.1.3)
    s = str(valor).strip()
    
    # 2. Remove tudo que não é número (letras, pontuação) (2.1.4)
    s_nums = re.sub(r'\D', '', s)
    
    if not s_nums:
        return ""
        
    # 3. Remove zeros à esquerda (2.1.1 e 2.1.2)
    # Obs: Se for cpf_convenio, a regra de negócio do arquivo original forçava 11 digitos, 
    # mas aqui estamos padronizando a limpeza base. O tratamento específico ocorre abaixo.
    return s_nums.lstrip('0')

def processar_importacao_lote(conn, df, table_name, mapping, import_id, file_path_original):
    cur = conn.cursor()
    try:
        erros = []
        df_proc = pd.DataFrame()
        cols_order = []
        
        table_full_name = f"banco_pf.{table_name}"
        
        cur.execute("UPDATE banco_pf.pf_historico_importacoes SET caminho_arquivo_original = %s WHERE id = %s", (file_path_original, import_id))

        cols_banco_raw = get_table_columns(table_name)
        cols_banco = [c[0] for c in cols_banco_raw]

        # =========================================================================
        # LÓGICA ESPECÍFICA 1: TELEFONES (Validação 9 Dígitos + Deduplicação)
        # =========================================================================
        if table_name == 'pf_telefones':
            col_cpf = next((k for k, v in mapping.items() if v == 'cpf'), None)
            col_whats = next((k for k, v in mapping.items() if v == 'tag_whats'), None)
            col_qualif = next((k for k, v in mapping.items() if v == 'tag_qualificacao'), None)
            map_tels = {k: v for k, v in mapping.items() if v and v.startswith('telefone_')}
            
            if not col_cpf: return 0, 0, ["Erro: Coluna 'CPF' é obrigatória."]
            
            new_rows = []
            for _, row in df.iterrows():
                # Aplica limpeza rigorosa no CPF
                cpf_val = row[col_cpf]
                cpf_limpo = limpar_cpf_regra_importacao(cpf_val)
                
                if not cpf_limpo: continue
                
                whats_val = str(row[col_whats]).upper().strip() if col_whats and pd.notna(row[col_whats]) else None
                qualif_val = str(row[col_qualif]).upper().strip() if col_qualif and pd.notna(row[col_qualif]) else None
                
                for col_origin, _ in map_tels.items():
                    tel_raw = row[col_origin]
                    if pd.notna(tel_raw):
                        tel_limpo = pf_core.limpar_apenas_numeros(tel_raw)
                        numero_final = None
                        
                        if len(tel_limpo) == 13 and tel_limpo.startswith("55"): 
                            tel_limpo = tel_limpo[2:] 
                        
                        if len(tel_limpo) == 11:
                            if tel_limpo[2] == '9':
                                numero_final = tel_limpo
                            else:
                                continue 
                        else:
                            continue
                        
                        if numero_final: 
                            row_dict = {
                                'cpf': cpf_limpo, 
                                'numero': numero_final, 
                                'tag_whats': whats_val, 
                                'tag_qualificacao': qualif_val, 
                                'data_atualizacao': datetime.now().strftime('%Y-%m-%d')
                            }
                            if 'importacao_id' in cols_banco:
                                row_dict['importacao_id'] = str(import_id)
                            new_rows.append(row_dict)
                            
            if not new_rows: return 0, 0, ["Nenhum telefone celular válido encontrado."]
            df_proc = pd.DataFrame(new_rows)
            df_proc.drop_duplicates(subset=['cpf', 'numero'], inplace=True)
            cols_order = list(df_proc.columns)

        # =========================================================================
        # LÓGICA GENÉRICA PARA DEMAIS TABELAS
        # =========================================================================
        else:
            df_proc = df.rename(columns=mapping)
            cols_permitidas = cols_banco + ['cpf', 'matricula', 'convenio']
            df_proc = df_proc[[c for c in df_proc.columns if c in cols_permitidas]]
            df_proc = df_proc.applymap(lambda x: str(x).upper().strip() if isinstance(x, str) else x)

            if 'importacao_id' in cols_banco:
                df_proc['importacao_id'] = str(import_id)
            
            if 'data_atualizacao' in cols_banco:
                df_proc['data_atualizacao'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # --- APLICAÇÃO DAS REGRAS DE CPF (2.1) ---
            for c in ['cpf']:
                if c in df_proc.columns:
                    # 1. Aplica a limpeza que remove zeros, espaços e letras
                    df_proc[c] = df_proc[c].apply(limpar_cpf_regra_importacao)
                    
                    # 2. Validação: Verifica se sobrou algum número válido
                    # Aceita se tiver dígitos suficientes. Como removemos zeros, um CPF 00123... vira "123..." (menos dígitos).
                    # A regra 2.1.2 diz "aceita". Então filtramos apenas vazios ou muito curtos (erro óbvio).
                    def validar_existencia_cpf(val): 
                        return len(str(val)) > 0 
                    
                    df_proc = df_proc[df_proc[c].apply(validar_existencia_cpf)]
                    
                    # Exceção específica: cpf_convenio pede zero a esquerda?
                    # O usuário pediu "2.1.1 sistema deve converter para número sem zero na frente".
                    # Mantenho a regra global sem zero. Se cpf_convenio precisar de padding visual, é feito na exportação.
                    if table_name == 'cpf_convenio':
                         # Caso legado exija zfill, descomentar abaixo. 
                         # Por hora, seguindo a regra 2.1.1 estrita: SEM ZERO NA FRENTE NO BANCO.
                         pass 

            if table_name == 'pf_emails' and 'cpf' in df_proc.columns and 'email' in df_proc.columns:
                 df_proc.drop_duplicates(subset=['cpf', 'email'], inplace=True)
            
            if table_name == 'pf_enderecos' and 'cpf' in df_proc.columns and 'cep' in df_proc.columns:
                 df_proc.drop_duplicates(subset=['cpf', 'cep'], inplace=True)

            cols_order = list(df_proc.columns)

        # ---------------------------------------------------------------------
        # CARGA NO BANCO (BULK VIA STAGING)
        # ---------------------------------------------------------------------
        staging_table = f"staging_import_{import_id}"
        cur.execute(f"CREATE TEMP TABLE {staging_table} (LIKE {table_full_name} INCLUDING DEFAULTS) ON COMMIT DROP")
        
        output = io.StringIO()
        df_proc.to_csv(output, sep='\t', header=False, index=False, na_rep='\\N')
        output.seek(0)
        cur.copy_expert(f"COPY {staging_table} ({', '.join(cols_order)}) FROM STDIN WITH CSV DELIMITER E'\t' NULL '\\N'", output)
        
        qtd_novos, qtd_atualizados = 0, 0
        
        pk_field = None
        unique_checks = [] 

        if table_name == 'pf_dados':
            pk_field = 'cpf'
        elif table_name == 'pf_telefones':
            unique_checks = ['cpf', 'numero']
        elif table_name == 'pf_emails':
            unique_checks = ['cpf', 'email']
        elif table_name == 'pf_enderecos':
            unique_checks = ['cpf', 'cep']
        elif table_name == 'pf_emprego_renda':
            pk_field = 'matricula'
        
        if pk_field:
            set_parts = []
            for c in cols_order:
                if c == pk_field: continue
                if c == 'importacao_id' and table_name == 'pf_dados':
                    expr = f"CASE WHEN t.importacao_id IS NULL OR t.importacao_id = '' THEN s.importacao_id::text ELSE t.importacao_id || ', ' || s.importacao_id::text END"
                    set_parts.append(f"{c} = {expr}")
                else:
                    set_parts.append(f"{c} = s.{c}")
            
            if set_parts:
                set_clause = ', '.join(set_parts)
                cur.execute(f"UPDATE {table_full_name} t SET {set_clause} FROM {staging_table} s WHERE t.{pk_field} = s.{pk_field}")
                qtd_atualizados = cur.rowcount
            
            cur.execute(f"INSERT INTO {table_full_name} ({', '.join(cols_order)}) SELECT {', '.join(cols_order)} FROM {staging_table} s WHERE NOT EXISTS (SELECT 1 FROM {table_full_name} t WHERE t.{pk_field} = s.{pk_field})")
            qtd_novos = cur.rowcount

        elif unique_checks:
            where_conditions = " AND ".join([f"t.{col} = s.{col}" for col in unique_checks])
            
            insert_query = f"""
                INSERT INTO {table_full_name} ({', '.join(cols_order)}) 
                SELECT {', '.join(cols_order)} 
                FROM {staging_table} s 
                WHERE NOT EXISTS (
                    SELECT 1 FROM {table_full_name} t WHERE {where_conditions}
                )
            """
            cur.execute(insert_query)
            qtd_novos = cur.rowcount

        else:
            cur.execute(f"INSERT INTO {table_full_name} ({', '.join(cols_order)}) SELECT {', '.join(cols_order)} FROM {staging_table} s")
            qtd_novos = cur.rowcount
            
            str_imp = str(import_id)
            if table_name in ['pf_telefones', 'pf_emails', 'pf_enderecos', 'pf_emprego_renda', 'cpf_convenio']:
                cur.execute(f"UPDATE banco_pf.pf_dados d SET importacao_id = CASE WHEN d.importacao_id IS NULL OR d.importacao_id = '' THEN %s ELSE d.importacao_id || ', ' || %s END FROM {staging_table} s WHERE d.cpf = s.cpf", (str_imp, str_imp))

        return qtd_novos, qtd_atualizados, erros

    except Exception as e: raise e

def interface_historico():
    st.markdown("### 📜 Histórico de Importações")
    if st.button("⬅️ Voltar"): st.session_state['import_step'] = 1; st.rerun()
    conn = pf_core.get_conn()
    if not conn: return
    try:
        df = pd.read_sql("SELECT * FROM banco_pf.pf_historico_importacoes ORDER BY id DESC", conn)
        conn.close()
        st.dataframe(df, hide_index=True)
    except: st.error("Erro histórico")

def interface_importacao():
    if st.session_state.get('import_step') == 'historico': interface_historico(); return

    c1, c2 = st.columns([1, 4])
    if c1.button("⬅️ Cancelar"): st.session_state.update({'pf_view': 'lista', 'import_step': 1}); st.rerun()
    if c2.button("📜 Histórico"): st.session_state['import_step'] = 'historico'; st.rerun()
    
    st.divider()
    
    mapa = {
        "Dados Cadastrais": "pf_dados",
        "Telefones": "pf_telefones",
        "E-mails": "pf_emails",
        "Endereços": "pf_enderecos",
        "Emprego e Renda": "pf_emprego_renda",
        "Contratos": "pf_contratos",
        "Contratos CLT": "pf_matricula_dados_clt",
        "CPF x Convênio": "cpf_convenio",
        "Convênio x Planilha": "convenio_por_planilha"
    }
    
    if st.session_state.get('import_step', 1) == 1:
        sel = st.selectbox("Tipo de Importação", list(mapa.keys()))
        st.session_state['import_table'] = mapa[sel]
        
        st.info("ℹ️ Aceita arquivos **.CSV** e **.XLSX (Excel)**.")
        st.warning("⚠️ **Regra de Importação:** Formato 'Geral' bloqueado. Use Texto ou Número.")
        
        st.markdown(f"###### 🗃️ Tabela SQL: `{mapa[sel]}` | Tipo: {sel}")
        
        uploaded = st.file_uploader("Selecione o arquivo", type=['csv', 'xlsx'])
        
        if uploaded:
            path = os.path.join(BASE_DIR_IMPORTS, f"{datetime.now().strftime('%Y%m%d%H%M')}_{uploaded.name}")
            with open(path, "wb") as f: f.write(uploaded.getbuffer())
            st.session_state['uploaded_file_path'] = path
            st.session_state['uploaded_file_name'] = uploaded.name
            
            df = None
            
            if uploaded.name.endswith('.xlsx'):
                with st.spinner("Validando formatação do Excel..."):
                    valido, msg_erro = validar_planilha_estrita(path)
                
                if not valido:
                    st.error(msg_erro)
                    try: os.remove(path)
                    except: pass
                    return
                
                try: df = pd.read_excel(path, dtype=str)
                except Exception as e: st.error(f"Erro ao ler Excel: {e}")
            else:
                try:
                    df = pd.read_csv(path, sep=';', encoding='utf-8', dtype=str)
                    if len(df.columns) <= 1: df = pd.read_csv(path, sep=',', encoding='utf-8', dtype=str)
                except:
                    try: df = pd.read_csv(path, sep=';', encoding='latin-1', dtype=str)
                    except: df = None
            
            if df is not None:
                st.session_state['import_df'] = df
                st.success(f"Arquivo aprovado! {len(df)} linhas encontradas.")
                if st.button("Avançar para Mapeamento"):
                    st.session_state['csv_map'] = {col: None for col in df.columns}
                    st.session_state['current_csv_idx'] = 0
                    st.session_state['import_step'] = 2
                    st.rerun()
            else: st.error("Falha na leitura do arquivo.")

    elif st.session_state['import_step'] == 2:
        st.markdown("### 🔗 Mapeamento de Colunas")
        df = st.session_state['import_df']
        cols_csv = list(df.columns)
        tbl = st.session_state['import_table']
        cols_db = [c[0] for c in get_table_columns(tbl) if c[0] not in ['id', 'data_criacao', 'importacao_id']]
        
        if tbl == 'pf_telefones' and 'cpf' not in cols_db: cols_db.insert(0, 'cpf')
        if tbl in ['pf_contratos', 'pf_matricula_dados_clt']:
             if 'matricula' in cols_db: cols_db.remove('matricula')
             cols_db = ['matricula', 'cpf (Gerar Matrícula)'] + cols_db

        c_l, c_r = st.columns([1, 2])
        with c_l:
            for idx, col in enumerate(cols_csv):
                mapped = st.session_state['csv_map'].get(col)
                status_icon = "❌" if mapped == "IGNORAR" else ("✅" if mapped else "❓")
                btn_label = f"{status_icon} {col} -> {mapped if mapped else '...'}"
                tipo_btn = "primary" if idx == st.session_state.get('current_csv_idx', 0) else "secondary"
                if st.button(btn_label, key=f"btn_col_{idx}", type=tipo_btn, use_container_width=True): 
                    st.session_state['current_csv_idx'] = idx; st.rerun()
        
        with c_r:
            curr_idx = st.session_state['current_csv_idx']
            col_atual = cols_csv[curr_idx]
            st.info(f"Mapeando coluna do Arquivo: **{col_atual}**")
            st.write("Amostra de dados:")
            st.code(df[col_atual].head(3).to_string(index=False))
            
            c_ig, c_prox = st.columns([1, 2])
            if c_ig.button("🚫 IGNORAR COLUNA", use_container_width=True): 
                st.session_state['csv_map'][col_atual] = "IGNORAR"
                if curr_idx < len(cols_csv)-1: st.session_state['current_csv_idx'] += 1
                st.rerun()
            
            cols_banco_cols = st.columns(3)
            for i, field in enumerate(cols_db):
                with cols_banco_cols[i % 3]:
                    if st.button(f"📥 {field}", key=f"map_to_{field}"):
                        st.session_state['csv_map'][col_atual] = 'cpf' if 'cpf' in field else ('matricula' if 'matricula' == field else field)
                        if 'Gerar Matrícula' in field: st.session_state['csv_map'][col_atual] = 'cpf' 
                        if curr_idx < len(cols_csv)-1: st.session_state['current_csv_idx'] += 1
                        st.rerun()

        st.divider()
        if st.button("🚀 INICIAR IMPORTAÇÃO", type="primary", use_container_width=True):
            conn = pf_core.get_conn()
            if conn:
                with st.spinner("Processando dados..."):
                    try:
                        cur = conn.cursor()
                        cur.execute("INSERT INTO banco_pf.pf_historico_importacoes (nome_arquivo) VALUES (%s) RETURNING id", (st.session_state['uploaded_file_name'],))
                        imp_id = cur.fetchone()[0]
                        conn.commit()
                        
                        mapping = {k: v for k, v in st.session_state['csv_map'].items() if v and v != "IGNORAR"}
                        res = processar_importacao_lote(conn, df, tbl, mapping, imp_id, st.session_state['uploaded_file_path'])
                        conn.commit()
                        
                        cur.execute("UPDATE banco_pf.pf_historico_importacoes SET qtd_novos=%s, qtd_atualizados=%s WHERE id=%s", (res[0], res[1], imp_id))
                        conn.commit(); conn.close()
                        
                        st.session_state['import_stats'] = res
                        st.session_state['import_step'] = 3; st.rerun()
                    except Exception as e: st.error(f"Erro: {e}")

    elif st.session_state['import_step'] == 3:
        st.balloons()
        st.success("✅ Importação Concluída!")
        res = st.session_state.get('import_stats', (0,0,[]))
        c1, c2 = st.columns(2)
        c1.metric("Novos", res[0])
        c2.metric("Atualizados", res[1])
        if res[2]: st.write(res[2])
        if st.button("Voltar"): st.session_state['import_step'] = 1; st.rerun()